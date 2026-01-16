from pathlib import Path
from typing import Dict, List

from fastapi import HTTPException, status
from openpyxl import load_workbook


def generate_preview(file_path: Path, max_rows: int = 50) -> Dict:
    """Generate preview data for all sheets of an Excel file."""
    workbook = None
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        if not workbook.sheetnames:
            return {"sheet": None, "sheets": []}

        sheets = []
        for sheet in workbook.worksheets:
            total_rows = sheet.max_row or 0
            limit = min(max_rows, total_rows)
            rows: List[List[str]] = []

            for row in sheet.iter_rows(min_row=1, max_row=limit, values_only=True):
                rows.append(["" if cell is None else str(cell) for cell in row])

            sheets.append(
                {
                    "name": sheet.title,
                    "rows": rows,
                    "total_rows": total_rows,
                }
            )

        return {"sheet": sheets[0] if sheets else None, "sheets": sheets}
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(exc)}"
        )
    finally:
        if workbook:
            workbook.close()
